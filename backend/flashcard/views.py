from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response

from .models import Deck, Tag, Card, UserFlashcardProgress
from .serializers import (
    DeckSerializer,
    TagSerializer,
    CardSerializer,
    UserFlashcardProgressSerializer,
    ExcelUploadSerializer
)
from openpyxl import load_workbook
from django.db import transaction
import io
import csv


class IsOwnerOrReadOnly(permissions.BasePermission):
    def has_object_permission(self, request, view, obj):
        if request.method in permissions.SAFE_METHODS:
            return True
        return getattr(obj, 'user', None) == request.user


class DeckViewSet(viewsets.ModelViewSet):
    queryset = Deck.objects.all()
    serializer_class = DeckSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrReadOnly]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def _read_excel_file(self, file_stream):
        """Reads an Excel file from a BytesIO stream and returns a list of rows."""
        wb = load_workbook(filename=file_stream, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Convert None to empty strings for consistency
            row_data = [str(cell) if cell is not None else '' for cell in row]
            # Consider ignoring completely empty rows if needed
            if any(val.strip() for val in row_data):
                rows.append(row_data)
        return rows

    def _read_csv_file(self, file_stream):
        """Reads a CSV file from a BytesIO stream and returns a list of rows."""
        content = file_stream.read()
        # You can add encoding detection if needed
        decoded = content.decode('utf-8')
        reader = csv.reader(io.StringIO(decoded))
        rows = [r for r in reader if any(c.strip() for c in r)]
        return rows

    def _detect_file_type_and_read(self, uploaded_file):
        """
        Attempt to read the file as Excel or CSV.
        You can improve by guessing based on filename or content.
        """
        # Simple heuristic: check extension
        file_name = getattr(uploaded_file, 'name', None) or ''
        ext = file_name.lower().split('.')[-1]
        file_stream = io.BytesIO(uploaded_file.read())

        if ext in ['xlsx', 'xls']:
            return self._read_excel_file(file_stream)
        elif ext in ['csv']:
            return self._read_csv_file(file_stream)
        else:
            # Default to Excel if uncertain
            return self._read_excel_file(file_stream)

    @action(detail=False, methods=['post'], url_path='import-excel')
    def upload_excel(self, request):
        """
        Step 1: Detect Columns
        POST /api/flashcard/decks/import-excel/
        - file: The uploaded file (Excel or CSV)

        Response:
          {
            "headers": [...],
            "preview": [... up to first 5 lines ...]
          }
        """
        serializer = ExcelUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        uploaded_file = serializer.validated_data['file']
        rows = self._detect_file_type_and_read(uploaded_file)

        if not rows:
            return Response({"error": "The uploaded file is empty or invalid."}, status=status.HTTP_400_BAD_REQUEST)

        headers = rows[0]
        preview = rows[1:6]  # Up to 5 lines preview
        return Response({
            "headers": headers,
            "preview": preview
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='confirm-import')
    def confirm_import(self, request):
        """
        Step 2: Confirm Import
        POST /api/flashcard/decks/confirm-import/

        Expected body:
        - file: the same file again or a stored reference
        - mapping: a dict { "ColumnName": "field_name", ... }
          field_name in: deck_title, front_text, back_text, tags

        Example:
        {
          "mapping": {
            "Deck Title": "deck_title",
            "Question": "front_text",
            "Answer": "back_text",
            "Labels": "tags"
          },
          "file": <the_file>,
        }

        The code will then create Deck and Card records accordingly.
        """
        serializer = ExcelUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        uploaded_file = serializer.validated_data['file']
        rows = self._detect_file_type_and_read(uploaded_file)
        if not rows:
            return Response({"error": "The file is empty."}, status=status.HTTP_400_BAD_REQUEST)

        mapping = request.data.get('mapping', {})
        if not mapping:
            return Response({"error": "No mapping provided."}, status=status.HTTP_400_BAD_REQUEST)

        headers = rows[0]
        data_rows = rows[1:]
        if not data_rows:
            return Response({"error": "No data rows found after headers."}, status=status.HTTP_400_BAD_REQUEST)

        # Reverse map: header -> index
        header_index = {h: i for i, h in enumerate(headers)}

        # Required fields to create a card
        required_fields = ['deck_title', 'front_text']
        # Check all required fields are mapped
        for rf in required_fields:
            if rf not in mapping.values():
                return Response({"error": f"Required field '{rf}' is not mapped."}, status=status.HTTP_400_BAD_REQUEST)

        # Identify column indexes
        deck_col = None
        front_col = None
        back_col = None
        tags_col = None

        for col_header, field_name in mapping.items():
            col_idx = header_index.get(col_header)
            if col_idx is None:
                # The user mapped a column that doesn't exist in headers
                continue
            if field_name == 'deck_title':
                deck_col = col_idx
            elif field_name == 'front_text':
                front_col = col_idx
            elif field_name == 'back_text':
                back_col = col_idx
            elif field_name == 'tags':
                tags_col = col_idx

        # If required columns are still None (maybe user didn't map correctly?)
        if deck_col is None or front_col is None:
            return Response({"error": "The required fields (deck_title, front_text) must be mapped correctly."},
                            status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        user = request.user
        with transaction.atomic():
            for row in data_rows:
                deck_title_val = row[deck_col].strip()
                front_text_val = row[front_col].strip() if front_col is not None else ''
                back_text_val = row[back_col].strip() if back_col is not None and row[back_col] else ''
                tags_val = row[tags_col].strip() if tags_col is not None and row[tags_col] else ''

                # Skip rows missing required fields
                if not deck_title_val or not front_text_val:
                    continue

                deck, _created = Deck.objects.get_or_create(
                    title=deck_title_val,
                    user=user,
                    defaults={'language': '', 'description': ''}
                )
                card = Card.objects.create(
                    deck=deck,
                    front_text=front_text_val,
                    back_text=back_text_val
                )

                if tags_val:
                    tag_names = [t.strip() for t in tags_val.split(',') if t.strip()]
                    for tag_name in tag_names:
                        tag, _ = Tag.objects.get_or_create(name=tag_name)
                        card.tags.add(tag)

                created_count += 1

        return Response(
            {"message": f"{created_count} cards created successfully."},
            status=status.HTTP_201_CREATED
        )


class TagViewSet(viewsets.ModelViewSet):
    queryset = Tag.objects.all()
    serializer_class = TagSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]


class CardViewSet(viewsets.ModelViewSet):
    queryset = Card.objects.all()
    serializer_class = CardSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrReadOnly]

    def get_queryset(self):
        return super().get_queryset()

    def perform_create(self, serializer):
        deck = serializer.validated_data.get('deck')
        if deck.user != self.request.user:
            raise PermissionError("Vous n'êtes pas propriétaire de ce deck.")
        serializer.save()


class UserFlashcardProgressViewSet(viewsets.ModelViewSet):
    queryset = UserFlashcardProgress.objects.all()
    serializer_class = UserFlashcardProgressSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrReadOnly]

    def get_queryset(self):
        return UserFlashcardProgress.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        card = serializer.validated_data.get('card')
        if card.deck.user != self.request.user:
            raise PermissionError("Vous ne pouvez suivre que vos propres cartes.")
        serializer.save(user=self.request.user)
