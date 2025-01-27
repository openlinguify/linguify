import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_unit_excel(output_file='units_template.xlsx'):
    # Données pour 10 unités
    sample_data = [
        {
            'id': 1,
            'title_en': 'Getting Started',
            'title_fr': 'Premiers Pas',
            'title_es': 'Primeros Pasos',
            'title_nl': 'Eerste Stappen',
            'description_en': 'Basic introductions, greetings, and essential phrases for beginners',
            'description_fr': 'Introductions basiques, salutations et phrases essentielles pour débutants',
            'description_es': 'Introducciones básicas, saludos y frases esenciales para principiantes',
            'description_nl': 'Basis introducties, begroetingen en essentiële zinnen voor beginners',
            'level': 'A1',
            'order': 1
        },
        {
            'id': 2,
            'title_en': 'Daily Routines',
            'title_fr': 'Routines Quotidiennes',
            'title_es': 'Rutinas Diarias',
            'title_nl': 'Dagelijkse Routines',
            'description_en': 'Learn to describe your daily activities and schedule',
            'description_fr': 'Apprenez à décrire vos activités quotidiennes et votre emploi du temps',
            'description_es': 'Aprende a describir tus actividades diarias y horario',
            'description_nl': 'Leer je dagelijkse activiteiten en schema beschrijven',
            'level': 'A1',
            'order': 2
        },
        {
            'id': 3,
            'title_en': 'Food and Dining',
            'title_fr': 'Alimentation et Restauration',
            'title_es': 'Comida y Restaurantes',
            'title_nl': 'Eten en Dineren',
            'description_en': 'Vocabulary and phrases for ordering food and dining out',
            'description_fr': 'Vocabulaire et phrases pour commander de la nourriture et manger au restaurant',
            'description_es': 'Vocabulario y frases para pedir comida y comer fuera',
            'description_nl': 'Woordenschat en zinnen voor het bestellen van eten en uit eten gaan',
            'level': 'A1',
            'order': 3
        },
        {
            'id': 4,
            'title_en': 'Travel and Transportation',
            'title_fr': 'Voyage et Transport',
            'title_es': 'Viajes y Transporte',
            'title_nl': 'Reizen en Transport',
            'description_en': 'Essential vocabulary for traveling and using public transportation',
            'description_fr': 'Vocabulaire essentiel pour voyager et utiliser les transports en commun',
            'description_es': 'Vocabulario esencial para viajar y usar el transporte público',
            'description_nl': 'Essentiële woordenschat voor reizen en openbaar vervoer',
            'level': 'A2',
            'order': 4
        },
        {
            'id': 5,
            'title_en': 'Shopping and Services',
            'title_fr': 'Shopping et Services',
            'title_es': 'Compras y Servicios',
            'title_nl': 'Winkelen en Diensten',
            'description_en': 'How to shop, ask for prices, and use various services',
            'description_fr': 'Comment faire des achats, demander les prix et utiliser divers services',
            'description_es': 'Cómo comprar, preguntar precios y usar varios servicios',
            'description_nl': 'Hoe te winkelen, prijzen vragen en verschillende diensten gebruiken',
            'level': 'A2',
            'order': 5
        },
        {
            'id': 6,
            'title_en': 'Health and Wellness',
            'title_fr': 'Santé et Bien-être',
            'title_es': 'Salud y Bienestar',
            'title_nl': 'Gezondheid en Welzijn',
            'description_en': 'Vocabulary for medical situations and discussing health',
            'description_fr': 'Vocabulaire pour les situations médicales et parler de la santé',
            'description_es': 'Vocabulario para situaciones médicas y hablar de salud',
            'description_nl': 'Woordenschat voor medische situaties en over gezondheid praten',
            'level': 'B1',
            'order': 6
        },
        {
            'id': 7,
            'title_en': 'Work and Career',
            'title_fr': 'Travail et Carrière',
            'title_es': 'Trabajo y Carrera',
            'title_nl': 'Werk en Carrière',
            'description_en': 'Professional vocabulary and workplace communication',
            'description_fr': 'Vocabulaire professionnel et communication au travail',
            'description_es': 'Vocabulario profesional y comunicación en el trabajo',
            'description_nl': 'Professionele woordenschat en communicatie op het werk',
            'level': 'B1',
            'order': 7
        },
        {
            'id': 8,
            'title_en': 'Culture and Entertainment',
            'title_fr': 'Culture et Divertissement',
            'title_es': 'Cultura y Entretenimiento',
            'title_nl': 'Cultuur en Entertainment',
            'description_en': 'Discussing arts, entertainment, and cultural activities',
            'description_fr': 'Discuter des arts, du divertissement et des activités culturelles',
            'description_es': 'Hablar de arte, entretenimiento y actividades culturales',
            'description_nl': 'Praten over kunst, entertainment en culturele activiteiten',
            'level': 'B2',
            'order': 8
        },
        {
            'id': 9,
            'title_en': 'Current Events and Media',
            'title_fr': 'Actualités et Médias',
            'title_es': 'Actualidad y Medios',
            'title_nl': 'Actualiteit en Media',
            'description_en': 'Discussing news, media, and current events',
            'description_fr': 'Discuter des nouvelles, des médias et de l\'actualité',
            'description_es': 'Hablar de noticias, medios y actualidad',
            'description_nl': 'Praten over nieuws, media en actualiteit',
            'level': 'C1',
            'order': 9
        },
        {
            'id': 10,
            'title_en': 'Advanced Communication',
            'title_fr': 'Communication Avancée',
            'title_es': 'Comunicación Avanzada',
            'title_nl': 'Geavanceerde Communicatie',
            'description_en': 'Complex discussions, debates, and advanced language skills',
            'description_fr': 'Discussions complexes, débats et compétences linguistiques avancées',
            'description_es': 'Discusiones complejas, debates y habilidades lingüísticas avanzadas',
            'description_nl': 'Complexe discussies, debatten en geavanceerde taalvaardigheden',
            'level': 'C1',
            'order': 10
        }
    ]

    # Créer le DataFrame
    df = pd.DataFrame(sample_data)

    # Créer un ExcelWriter
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Convertir le DataFrame en Excel
        df.to_excel(writer, sheet_name='Units', index=False)
        
        # Accéder au classeur et à la feuille
        workbook = writer.book
        worksheet = writer.sheets['Units']

        # Styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Ajouter une ligne de titre
        worksheet.insert_rows(1)
        title_cell = worksheet.cell(row=1, column=1, value='Language Learning Units')
        title_cell.font = Font(name='Arial', size=14, bold=True)
        worksheet.merge_cells(f'A1:{get_column_letter(worksheet.max_column)}1')
        title_cell.alignment = Alignment(horizontal='center')

        # Formater les en-têtes
        for cell in worksheet[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Formater les cellules de données
        for row in worksheet.iter_rows(min_row=3):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Définir la largeur des colonnes optimale
        column_widths = {
            'id': 8,
            'title': 30,
            'description': 50,
            'level': 10,
            'order': 10
        }

        # Ajuster les colonnes
        for idx, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(idx)
            if 'title' in column:
                worksheet.column_dimensions[column_letter].width = column_widths['title']
            elif 'description' in column:
                worksheet.column_dimensions[column_letter].width = column_widths['description']
            elif column in column_widths:
                worksheet.column_dimensions[column_letter].width = column_widths[column]

        # Figer les volets
        worksheet.freeze_panes = 'A3'

        # Ajouter des filtres
        worksheet.auto_filter.ref = f'A2:{get_column_letter(worksheet.max_column)}{worksheet.max_row}'

        # Validation pour le niveau
        level_range = ['A1', 'A2', 'B1', 'B2', 'C1', 'C2']
        level_col = df.columns.get_loc('level') + 1
        for row in range(3, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=level_col)
            cell.data_validation = openpyxl.worksheet.datavalidation.DataValidation(
                type='list',
                formula1=f'"{",".join(level_range)}"',
                allow_blank=True
            )

        # Sauvegarder
        workbook.save(output_file)

    print(f'Fichier Excel créé avec succès : {output_file}')

if __name__ == '__main__':
    create_unit_excel()