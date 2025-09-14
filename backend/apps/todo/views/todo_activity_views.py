
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Avg, Max
from django.utils import timezone
from datetime import timedelta
from django.template.loader import render_to_string
from django.views.decorators.http import require_http_methods
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
import logging
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404
from ..models.todo_models import Task, Project, PersonalStageType, Tag, Category, Note, Reminder, TaskTemplate

logger = logging.getLogger(__name__)

# HTMX Helper Mixin
class HTMXResponseMixin:
    """Mixin for handling HTMX requests and responses"""
    
    def get_htmx_template(self, template_name=None):
        """Get template for HTMX partial response"""
        if template_name:
            return template_name
        return getattr(self, 'htmx_template_name', self.template_name)
    
    def render_htmx_response(self, context, template_name=None):
        """Render partial template for HTMX"""
        template = self.get_htmx_template(template_name)
        html = render_to_string(template, context, request=self.request)
        return HttpResponse(html)
    
    def is_htmx(self):
        """Check if request is from HTMX"""
        return self.request.headers.get('HX-Request') == 'true'

class TodoActivityView(LoginRequiredMixin, HTMXResponseMixin, TemplateView):
    """HTMX-powered Activity/Timeline view - openlinguify-style activity tracking"""
    template_name = 'todo/views/activity.html'
    htmx_template_name = 'todo/partials/activity_feed.html'
    
    def get(self, request, *args, **kwargs):
        # If HTMX request, return partial content
        if self.is_htmx():
            return self.render_htmx_response(self.get_context_data())
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Get recent activity (last 30 days)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        
        # Recently created tasks
        recent_tasks = Task.objects.filter(
            user=user,
            created_at__gte=thirty_days_ago
        ).select_related('project', 'personal_stage_type').prefetch_related('tags').order_by('-created_at')[:20]
        
        # Recently completed tasks
        completed_tasks = Task.objects.filter(
            user=user,
            state='1_done',
            completed_at__gte=thirty_days_ago
        ).select_related('project', 'personal_stage_type').prefetch_related('tags').order_by('-completed_at')[:20]
        
        # Overdue tasks
        overdue_tasks = Task.objects.filter(
            user=user,
            due_date__lt=timezone.now(),
            state__in=['1_todo', '1_in_progress'],
            active=True
        ).select_related('project', 'personal_stage_type').prefetch_related('tags').order_by('due_date')
        
        # Due today
        today = timezone.now().date()
        due_today_tasks = Task.objects.filter(
            user=user,
            due_date__date=today,
            state__in=['1_todo', '1_in_progress'],
            active=True
        ).select_related('project', 'personal_stage_type').prefetch_related('tags').order_by('due_date')
        
        # Due this week  
        week_end = today + timedelta(days=7)
        week_start = today - timedelta(days=today.weekday())
        
        due_week_tasks = Task.objects.filter(
            user=user,
            due_date__date__range=[today, week_end],
            state__in=['1_todo', '1_in_progress'],
            active=True
        ).exclude(due_date__date=today).select_related('project', 'personal_stage_type').prefetch_related('tags').order_by('due_date')
        
        # Enhanced activity statistics with dashboard metrics
        stats = {
            'tasks_created_today': Task.objects.filter(
                user=user,
                created_at__date=today
            ).count(),
            'tasks_completed_today': Task.objects.filter(
                user=user,
                completed_at__date=today
            ).count(),
            'tasks_created_this_week': Task.objects.filter(
                user=user,
                created_at__date__gte=week_start
            ).count(),
            'tasks_completed_this_week': Task.objects.filter(
                user=user,
                completed_at__date__gte=week_start
            ).count(),
            'total_active_tasks': Task.objects.filter(
                user=user,
                active=True,
                state__in=['1_todo', '1_in_progress']
            ).count(),
            'overdue_count': overdue_tasks.count(),
            'due_today_count': due_today_tasks.count(),
            'due_week_count': due_week_tasks.count(),
        }
        
        # Calculate completion rate
        if stats['tasks_created_this_week'] > 0:
            stats['completion_rate'] = round(
                (stats['tasks_completed_this_week'] / stats['tasks_created_this_week']) * 100, 1
            )
        else:
            stats['completion_rate'] = 0
            
        # Calculate productivity score
        stats['productivity_score'] = round(stats['tasks_completed_this_week'] / 7, 1)
        
        context.update({
            'recent_tasks': recent_tasks,
            'completed_tasks': completed_tasks,
            'overdue_tasks': overdue_tasks,
            'due_today_tasks': due_today_tasks,
            'due_week_tasks': due_week_tasks,
            'activity_stats': stats,
            'view_mode': 'activity',
            'today': today,
        })
        
        return context




class ActivityTimelineHTMXView(LoginRequiredMixin, HTMXResponseMixin, TemplateView):
    """HTMX endpoint for activity timeline"""
    template_name = 'todo/partials/activity_timeline.html'
    
    def get(self, request, *args, **kwargs):
        user = request.user
        
        # Get pagination params
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        
        # Get filter params
        activity_type = request.GET.get('type', 'all')  # all, tasks, projects, notes
        time_range = int(request.GET.get('range', 7))  # days
        
        start_date = timezone.now() - timedelta(days=time_range)
        
        activities = []
        
        if activity_type in ['all', 'tasks']:
            # Task activities
            completed_tasks = Task.objects.filter(
                user=user,
                completed_at__gte=start_date
            ).select_related('project', 'personal_stage_type').order_by('-completed_at')
            
            for task in completed_tasks:
                activities.append({
                    'type': 'task_completed',
                    'timestamp': task.completed_at,
                    'task': task,
                    'icon': 'bi-check-circle-fill',
                    'color': 'text-green-600',
                    'title': 'Task Completed',
                    'description': task.title,
                    'project': task.project
                })
            
            created_tasks = Task.objects.filter(
                user=user,
                created_at__gte=start_date
            ).select_related('project', 'personal_stage_type').order_by('-created_at')
            
            for task in created_tasks:
                activities.append({
                    'type': 'task_created',
                    'timestamp': task.created_at,
                    'task': task,
                    'icon': 'bi-plus-circle-fill',
                    'color': 'text-blue-600',
                    'title': 'Task Created',
                    'description': task.title,
                    'project': task.project
                })
        
        if activity_type in ['all', 'projects']:
            # Project activities
            projects = Project.objects.filter(
                user=user,
                created_at__gte=start_date
            ).order_by('-created_at')
            
            for project in projects:
                activities.append({
                    'type': 'project_created',
                    'timestamp': project.created_at,
                    'project': project,
                    'icon': 'bi-folder-plus',
                    'color': 'text-purple-600',
                    'title': 'Project Created',
                    'description': project.name
                })
        
        # Sort by timestamp
        activities.sort(key=lambda x: x['timestamp'], reverse=True)
        
        # Pagination
        start = (page - 1) * per_page
        end = start + per_page
        paginated_activities = activities[start:end]
        
        has_more = len(activities) > end
        
        context = {
            'activities': paginated_activities,
            'has_more': has_more,
            'next_page': page + 1,
            'activity_type': activity_type,
            'time_range': time_range,
        }
        
        return self.render_to_response(context)


class ActivityStatsHTMXView(LoginRequiredMixin, HTMXResponseMixin, TemplateView):
    """HTMX endpoint for activity statistics"""
    template_name = 'todo/partials/activity_stats.html'
    
    def get(self, request, *args, **kwargs):
        user = request.user
        time_range = int(request.GET.get('range', 7))  # days
        
        start_date = timezone.now() - timedelta(days=time_range)
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())
        
        # Calculate statistics
        stats = {
            'tasks_created_today': Task.objects.filter(
                user=user,
                created_at__date=today
            ).count(),
            'tasks_completed_today': Task.objects.filter(
                user=user,
                completed_at__date=today
            ).count(),
            'tasks_created_this_week': Task.objects.filter(
                user=user,
                created_at__date__gte=week_start
            ).count(),
            'tasks_completed_this_week': Task.objects.filter(
                user=user,
                completed_at__date__gte=week_start
            ).count(),
            'tasks_created': Task.objects.filter(
                user=user,
                created_at__gte=start_date
            ).count(),
            'tasks_completed': Task.objects.filter(
                user=user,
                completed_at__gte=start_date
            ).count(),
        }
        
        # Calculate trends (compare with previous period)
        prev_start = start_date - timedelta(days=time_range)
        prev_stats = {
            'tasks_created': Task.objects.filter(
                user=user,
                created_at__gte=prev_start,
                created_at__lt=start_date
            ).count(),
            'tasks_completed': Task.objects.filter(
                user=user,
                completed_at__gte=prev_start,
                completed_at__lt=start_date
            ).count(),
            'projects_created': Project.objects.filter(
                user=user,
                created_at__gte=prev_start,
                created_at__lt=start_date
            ).count(),
        }
        
        # Calculate percentage changes
        trends = {}
        for key in stats:
            if key in prev_stats:
                if prev_stats[key] > 0:
                    change = ((stats[key] - prev_stats[key]) / prev_stats[key]) * 100
                    trends[key] = {
                        'value': round(change, 1),
                        'direction': 'up' if change > 0 else 'down' if change < 0 else 'same',
                        'color': 'text-green-600' if change > 0 else 'text-red-600' if change < 0 else 'text-gray-600'
                    }
                else:
                    trends[key] = {
                        'value': 100 if stats[key] > 0 else 0,
                        'direction': 'up' if stats[key] > 0 else 'same',
                        'color': 'text-green-600' if stats[key] > 0 else 'text-gray-600'
                    }
        
        # Productivity metrics
        avg_daily_tasks = stats['tasks_completed'] / max(time_range, 1)
        completion_rate = (stats['tasks_completed'] / max(stats['tasks_created'], 1)) * 100 if stats['tasks_created'] > 0 else 0
        
        productivity = {
            'avg_daily_tasks': round(avg_daily_tasks, 1),
            'completion_rate': round(completion_rate, 1),
            'productivity_score': min(100, round(avg_daily_tasks * 10, 0)),  # Simple scoring system
        }
        
        context = {
            'stats': stats,
            'trends': trends,
            'productivity': productivity,
            'time_range': time_range,
        }
        
        return self.render_to_response(context)


class ActivityExportView(LoginRequiredMixin, TemplateView):
    """Export activity data in various formats"""
    
    def get(self, request, *args, **kwargs):
        user = request.user
        export_format = request.GET.get('format', 'json')  # json, csv, xlsx
        time_range = int(request.GET.get('range', 30))  # days
        
        start_date = timezone.now() - timedelta(days=time_range)
        
        # Get user export settings from database
        from ..models import TodoSettings
        try:
            settings = TodoSettings.objects.get(user=user)
            include_completed = settings.include_completed_in_exports
        except TodoSettings.DoesNotExist:
            # Default to including completed tasks if no settings found
            include_completed = True
        
        # Collect activity data
        activities = []
        
        # Task activities - filter based on settings
        tasks_query = Task.objects.filter(
            user=user,
            created_at__gte=start_date
        )
        
        # Exclude completed tasks if setting is disabled
        if not include_completed:
            tasks_query = tasks_query.exclude(state='1_done')
        
        tasks = tasks_query.select_related('project', 'personal_stage_type').order_by('-created_at')
        
        for task in tasks:
            # Get priority display name
            if hasattr(task, 'priority') and task.priority is not None:
                if task.priority == 0:
                    priority_display = 'Normal'
                elif task.priority == 1:
                    priority_display = 'High'
                elif task.priority == 2:
                    priority_display = 'Critical'
                else:
                    priority_display = 'Normal'
            else:
                priority_display = 'Normal'
            
            # Get state display name
            state_display = ''
            if hasattr(task, 'state'):
                if task.state == '0_todo':
                    state_display = 'To Do'
                elif task.state == '1_progress':
                    state_display = 'In Progress'
                elif task.state == '2_review':
                    state_display = 'Review'
                elif task.state == '1_done':
                    state_display = 'Done'
                else:
                    state_display = task.personal_stage_type.name if task.personal_stage_type else 'Unknown'
            
            # Get tags
            tags = ', '.join([tag.name for tag in task.tags.all()]) if hasattr(task, 'tags') else ''
            
            activities.append({
                'id': task.id,
                'title': task.title,
                'description': task.description[:100] + '...' if task.description and len(task.description) > 100 else task.description or '',
                'state': state_display,
                'priority': priority_display,
                'project': task.project.name if task.project else 'No Project',
                'stage': task.personal_stage_type.name if task.personal_stage_type else 'No Stage',
                'tags': tags,
                'created_at': task.created_at.isoformat(),
                'updated_at': task.updated_at.isoformat() if hasattr(task, 'updated_at') and task.updated_at else '',
                'completed_at': task.completed_at.isoformat() if task.completed_at else '',
                'due_date': task.due_date.isoformat() if task.due_date else '',
                'progress_percentage': task.progress_percentage if hasattr(task, 'progress_percentage') else 0,
                'is_overdue': 'Yes' if task.due_date and task.due_date < timezone.now() and task.state != '1_done' else 'No',
                'active': 'Yes' if task.active else 'No',
            })
        
        if export_format == 'json':
            return self._export_json(activities)
        elif export_format == 'csv':
            return self._export_csv(activities)
        elif export_format == 'xlsx':
            return self._export_xlsx(activities)
        else:
            return JsonResponse({'error': 'Unsupported format'}, status=400)
    
    def _export_json(self, activities):
        """Export as JSON"""
        response = JsonResponse({'activities': activities})
        response['Content-Disposition'] = 'attachment; filename="activity_export.json"'
        return response
    
    def _export_csv(self, activities):
        """Export as CSV with UTF-8 encoding"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="tasks_export.csv"'
        # Add BOM for Excel compatibility with UTF-8
        response.write('\ufeff')
        
        if activities:
            writer = csv.DictWriter(response, fieldnames=activities[0].keys(), lineterminator='\n')
            writer.writeheader()
            for activity in activities:
                # Ensure all values are properly encoded strings
                encoded_activity = {}
                for key, value in activity.items():
                    if value is None:
                        encoded_activity[key] = ''
                    else:
                        encoded_activity[key] = str(value)
                writer.writerow(encoded_activity)
        
        return response
    
    def _export_xlsx(self, activities):
        """Export as Excel file"""
        try:
            import openpyxl
            from django.http import HttpResponse
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Activity Export"
            
            if activities:
                # Write headers
                headers = list(activities[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Write data
                for row, activity in enumerate(activities, 2):
                    for col, key in enumerate(headers, 1):
                        value = activity[key]
                        # Convert UUID objects to strings for Excel compatibility
                        if hasattr(value, 'hex'):  # Check if it's a UUID
                            value = str(value)
                        ws.cell(row=row, column=col, value=value)
            
            # Save to BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="tasks_export.xlsx"'
            return response
            
        except ImportError:
            return JsonResponse({'error': 'Excel export not available (openpyxl not installed)'}, status=400)
